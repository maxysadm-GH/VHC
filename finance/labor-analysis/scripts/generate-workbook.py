#!/usr/bin/env python3
"""
generate-workbook.py
Rebuilds the WIP/FG Labor Analysis workbook with formula-based calculations

Usage:
    python generate-workbook.py --payroll classified-payroll.csv --kpi kpi-data.csv --output WIP_FG_Analysis.xlsx
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from datetime import datetime
import argparse


# Styles
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
HDR_FILL = PatternFill("solid", fgColor="2F5496")
TITLE_FONT = Font(bold=True, size=18, color="1F4E79")
SECTION_FONT = Font(bold=True, size=14, color="1F4E79")
FORMULA_FONT = Font(italic=True, size=9, color="0070C0")
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
THICK_BORDER = Border(
    left=Side(style='medium'),
    right=Side(style='medium'),
    top=Side(style='medium'),
    bottom=Side(style='medium')
)

OPS_FILL = PatternFill("solid", fgColor="4472C4")
SGA_FILL = PatternFill("solid", fgColor="70AD47")
RETAIL_FILL = PatternFill("solid", fgColor="FFC000")
WIP_FILL = PatternFill("solid", fgColor="BDD7EE")
FG_FILL = PatternFill("solid", fgColor="C5E0B3")
GRAY_FILL = PatternFill("solid", fgColor="D9D9D9")


def create_raw_pay_tab(wb, df):
    """Create Tab 11 - Raw Pay with classified payroll data"""
    ws = wb.create_sheet('11-Raw Pay')
    
    # Title
    ws.cell(row=1, column=1, value="RAW: PAYROLL (Dual Classification)").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Records: {len(df)}").font = Font(italic=True, size=10, color="666666")
    
    # Headers at row 3
    headers = ['Pay_Date', 'Department', 'Orig_4Cat', 'J_CostGroup', 'J_SubDept', 'J_OpsType', 'Total_Earnings']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.border = BORDER
    
    # Data starting row 4
    row = 4
    for _, record in df.iterrows():
        ws.cell(row=row, column=1, value=record.get('Pay_Date'))
        ws.cell(row=row, column=2, value=record.get('Department'))
        ws.cell(row=row, column=3, value=record.get('Orig_4Cat'))
        ws.cell(row=row, column=4, value=record.get('J_CostGroup'))
        ws.cell(row=row, column=5, value=record.get('J_SubDept'))
        ws.cell(row=row, column=6, value=record.get('J_OpsType'))
        ws.cell(row=row, column=7, value=record.get('Total_Earnings'))
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 14
    
    return row - 1  # Return last data row


def create_raw_cu_tab(wb, df):
    """Create Tab 10 - Raw CU with KPI/production data"""
    ws = wb.create_sheet('10-Raw CU')
    
    # Title
    ws.cell(row=1, column=1, value="RAW: KPI PRODUCTION DATA").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Records: {len(df)}").font = Font(italic=True, size=10, color="666666")
    
    # Headers at row 3
    headers = ['Date', 'Part_Num', 'Part_Prefix', 'Product_Type', 'KPI_Units', 'Source']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.border = BORDER
    
    # Data starting row 4
    row = 4
    for _, record in df.iterrows():
        ws.cell(row=row, column=1, value=record.get('Date'))
        ws.cell(row=row, column=2, value=str(record.get('Part_Num', '')))
        ws.cell(row=row, column=3, value=str(record.get('Part_Prefix', '')))
        ws.cell(row=row, column=4, value=record.get('Product_Type'))
        ws.cell(row=row, column=5, value=record.get('KPI_Units'))
        ws.cell(row=row, column=6, value=record.get('Source', 'ClickUp'))
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 12
    
    return row - 1  # Return last data row


def create_executive_summary(wb, pay_end, cu_end):
    """Create Tab 1 - Executive Summary with formulas"""
    ws = wb.create_sheet('1-Executive Summary', 0)
    
    PAY_SHEET = "'11-Raw Pay'"
    CU_SHEET = "'10-Raw CU'"
    
    row = 1
    
    # Title
    ws.cell(row=row, column=1, value="2025 LABOR COST ANALYSIS").font = TITLE_FONT
    row += 1
    ws.cell(row=row, column=1, value="Executive Summary - J's 3-Tier Structure").font = Font(italic=True, size=10, color="666666")
    row += 3
    
    # Section: Labor by Cost Group
    ws.cell(row=row, column=1, value="LABOR BY COST GROUP").font = SECTION_FONT
    row += 2
    
    # Headers
    headers = ['Cost Group', 'Annual Cost', '% of Total']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.border = BORDER
    row += 1
    data_start = row
    
    # OPERATIONS
    ws.cell(row=row, column=1, value='OPERATIONS').fill = OPS_FILL
    ws.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=row, column=1).border = BORDER
    ws.cell(row=row, column=2, value=f'=SUMIF({PAY_SHEET}!$D$4:$D${pay_end},"operations",{PAY_SHEET}!$G$4:$G${pay_end})')
    ws.cell(row=row, column=2).number_format = '$#,##0'
    ws.cell(row=row, column=2).border = BORDER
    ws.cell(row=row, column=3, value=f'=B{row}/B${data_start+3}')
    ws.cell(row=row, column=3).number_format = '0.0%'
    ws.cell(row=row, column=3).border = BORDER
    row += 1
    
    # SG&A
    ws.cell(row=row, column=1, value='SG&A').fill = SGA_FILL
    ws.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=row, column=1).border = BORDER
    ws.cell(row=row, column=2, value=f'=SUMIF({PAY_SHEET}!$D$4:$D${pay_end},"sg&a",{PAY_SHEET}!$G$4:$G${pay_end})')
    ws.cell(row=row, column=2).number_format = '$#,##0'
    ws.cell(row=row, column=2).border = BORDER
    ws.cell(row=row, column=3, value=f'=B{row}/B${data_start+3}')
    ws.cell(row=row, column=3).number_format = '0.0%'
    ws.cell(row=row, column=3).border = BORDER
    row += 1
    
    # RETAIL
    ws.cell(row=row, column=1, value='RETAIL').fill = RETAIL_FILL
    ws.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=row, column=1).border = BORDER
    ws.cell(row=row, column=2, value=f'=SUMIF({PAY_SHEET}!$D$4:$D${pay_end},"retail",{PAY_SHEET}!$G$4:$G${pay_end})')
    ws.cell(row=row, column=2).number_format = '$#,##0'
    ws.cell(row=row, column=2).border = BORDER
    ws.cell(row=row, column=3, value=f'=B{row}/B${data_start+3}')
    ws.cell(row=row, column=3).number_format = '0.0%'
    ws.cell(row=row, column=3).border = BORDER
    data_end = row
    row += 1
    
    # TOTAL
    ws.cell(row=row, column=1, value='TOTAL').font = Font(bold=True)
    ws.cell(row=row, column=1).fill = GRAY_FILL
    ws.cell(row=row, column=1).border = THICK_BORDER
    ws.cell(row=row, column=2, value=f'=SUM(B{data_start}:B{data_end})')
    ws.cell(row=row, column=2).number_format = '$#,##0'
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=2).fill = GRAY_FILL
    ws.cell(row=row, column=2).border = THICK_BORDER
    ws.cell(row=row, column=3, value='=1')
    ws.cell(row=row, column=3).number_format = '0.0%'
    ws.cell(row=row, column=3).fill = GRAY_FILL
    ws.cell(row=row, column=3).border = THICK_BORDER
    
    # Column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 14
    
    return ws


def generate_workbook(payroll_file, kpi_file, output_file):
    """Generate complete workbook with formulas"""
    
    # Load data
    pay_df = pd.read_csv(payroll_file)
    kpi_df = pd.read_csv(kpi_file)
    
    print(f"Loaded {len(pay_df)} payroll records")
    print(f"Loaded {len(kpi_df)} KPI records")
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create raw data tabs first (to get row counts)
    pay_end = create_raw_pay_tab(wb, pay_df)
    cu_end = create_raw_cu_tab(wb, kpi_df)
    
    print(f"Payroll data: rows 4-{pay_end}")
    print(f"KPI data: rows 4-{cu_end}")
    
    # Create summary tabs with formulas
    create_executive_summary(wb, pay_end, cu_end)
    
    # Save
    wb.save(output_file)
    print(f"\n✓ Saved workbook to {output_file}")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Generate labor analysis workbook')
    parser.add_argument('--payroll', '-p', required=True, help='Classified payroll CSV')
    parser.add_argument('--kpi', '-k', required=True, help='KPI/production CSV')
    parser.add_argument('--output', '-o', required=True, help='Output Excel file')
    
    args = parser.parse_args()
    generate_workbook(args.payroll, args.kpi, args.output)
